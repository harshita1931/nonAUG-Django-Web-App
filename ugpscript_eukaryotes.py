#!/usr/bin/python

import sqlite3
import openpyxl
import sys
import csv

conn = sqlite3.connect('db.sqlite3')	#confirm the name
print ("opened database successfully")

wb = openpyxl.load_workbook('eukaryotes.xlsx')
sheet = wb.get_sheet_by_name('all')

#print (sheet)

inp = []
temp = []

for rows in range(2, 213):
	sno = rows-1
	temp.append(sno)
	geneid = sheet.cell(row = rows, column = 1).value
	temp.append(geneid)
	mRNAid = sheet.cell(row = rows, column = 2).value
	temp.append(mRNAid)
	protid = sheet.cell(row = rows, column = 3).value
	temp.append(protid)
	org = sheet.cell(row = rows, column = 4).value
	temp.append(org)
	gene_name = sheet.cell(row = rows, column = 5).value
	temp.append(gene_name)
	gene_desc = sheet.cell(row = rows, column = 6).value
	temp.append(gene_desc)
	prot_desc = sheet.cell(row = rows, column = 7).value
	temp.append(prot_desc)
	transcript = sheet.cell(row = rows, column = 8).value
	temp.append(transcript)
	chrom = sheet.cell(row = rows, column = 9).value
	temp.append(chrom)
	transcript_start = sheet.cell(row = rows, column = 10).value
	temp.append(transcript_start)
	transcript_end = sheet.cell(row = rows, column = 11).value
	temp.append(transcript_end)
	strand = sheet.cell(row = rows, column = 12).value
	temp.append(strand)
	codon = sheet.cell(row = rows, column = 13).value
	temp.append(codon)
	upstreamATG = sheet.cell(row = rows, column = 14).value
	temp.append(upstreamATG)
	prot_func = sheet.cell(row = rows, column = 15).value
	temp.append(prot_func)
	uniprot_swissprot_ID = sheet.cell(row = rows, column = 16).value
	temp.append(uniprot_swissprot_ID)
	pdbID = sheet.cell(row = rows, column = 17).value
	temp.append(pdbID)
	bio_proc = sheet.cell(row = rows, column = 18).value
	temp.append(bio_proc)
	cell_comp = sheet.cell(row = rows, column = 19).value
	temp.append(cell_comp)
	mol_func = sheet.cell(row = rows, column = 20).value
	temp.append(mol_func)
	 
	inp.append(temp)
	temp = []


conn.executemany("INSERT INTO ugp_eukaryotestablenew2 VALUES (?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?,?,?,?,?,?,?,?,?)", inp)	

conn.commit()

print ("Records created successfully")
conn.close()
