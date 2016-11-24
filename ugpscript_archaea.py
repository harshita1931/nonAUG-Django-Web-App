#!/usr/bin/python

import sqlite3
import openpyxl
import sys
import csv

conn = sqlite3.connect('db.sqlite3')
print ("opened database successfully")

wb = openpyxl.load_workbook('archaea.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


inp = []
temp = []

for rows in range(1, 42832):	
	sno = rows
	temp.append(sno)
	chrom_acc_num = sheet.cell(row = rows, column = 1).value
	temp.append(chrom_acc_num)
	organism =  sheet.cell(row = rows, column = 2).value
	temp.append(organism)
	CDSstart = sheet.cell(row = rows, column = 3).value
	temp.append(CDSstart)
	CDSend = sheet.cell(row = rows, column = 4).value
	temp.append(CDSend)
	strand = sheet.cell(row = rows, column = 5).value
	temp.append(strand)
	startcodon = sheet.cell(row = rows, column = 6).value
	temp.append(startcodon)
	prot_acc = sheet.cell(row = rows, column = 7).value
	temp.append(prot_acc)
	prot_description = sheet.cell(row = rows, column = 8).value
	temp.append(prot_description)
	locus_tag = sheet.cell(row = rows, column = 9).value
	temp.append(locus_tag)
	sequence = sheet.cell(row = rows, column = 10).value
	temp.append(sequence)
	inp.append(temp)
	temp = []

conn.executemany("INSERT INTO ugp_archaeatable VALUES (?,?, ?, ?, ?, ?, ?, ?, ?, ?,?)", inp)		
conn.commit()

print ("Records created successfully")
conn.close()
